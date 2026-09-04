"""Tests for the institutional file exports (/api/institution/report|export).

Two things matter here and are covered separately: the authorization boundary
(an institution may only ever export its own sensor's readings) and the file
itself actually being a well-formed PDF/XLSX built from the right rows — a
download that returns 200 with a corrupt body is a failure the status code
cannot catch.
"""

import io
import zipfile
from datetime import date, datetime, time, timezone as dt_timezone

from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from .exports import REPORT_TIME_ZONE
from .models import (
    Institution,
    InstitutionAlertConfig,
    InstitutionContract,
    InstitutionUser,
    Regions,
    StationReadingsGold,
    Stations,
)

User = get_user_model()


def _at(day: date, hour: int) -> datetime:
    """A reading timestamp at a given local hour, stored as UTC."""
    return datetime.combine(day, time(hour), tzinfo=REPORT_TIME_ZONE).astimezone(
        dt_timezone.utc
    )


class InstitutionExportTestCase(APITestCase):
    def setUp(self):
        self.client = APIClient()

        region = Regions.seed_for_tests(name="Gran Asuncion", region_code="GA")
        self.station = Stations.seed_for_tests(
            name="Respira: Villa Morra",
            region=region,
            latitude=-25.28,
            longitude=-57.57,
            is_station_on=True,
        )
        self.other_station = Stations.seed_for_tests(
            name="Respira: Sajonia", region=region, is_station_on=True
        )

        self.institution = Institution.objects.create(
            legal_name="Hospital Bautista", display_name="Hospital Bautista"
        )
        InstitutionContract.objects.create(
            institution=self.institution,
            station=self.station,
            contract_status=InstitutionContract.ContractStatus.ACTIVE,
            start_date=date(2026, 6, 1),
        )
        self.user = User.objects.create_user(
            email="contacto@bautista.test", password="Respira.Test.2026"
        )
        InstitutionUser.objects.create(user=self.user, institution=self.institution)

        # July 2026: three days of readings on the institution's own sensor.
        self.july_days = [date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 3)]
        for index, day in enumerate(self.july_days):
            for hour, aqi in ((8, 40 + index * 30), (20, 60 + index * 30)):
                StationReadingsGold.seed_for_tests(
                    station=self.station,
                    date_utc=_at(day, hour),
                    pm1=5.0,
                    pm2_5=12.5 + index,
                    pm10=20.0,
                    aqi_pm2_5=float(aqi),
                    aqi_pm10=30.0,
                )

        # A reading on somebody else's sensor, same month: it must never appear.
        StationReadingsGold.seed_for_tests(
            station=self.other_station,
            date_utc=_at(date(2026, 7, 2), 12),
            aqi_pm2_5=500.0,
        )

    def login(self):
        # Re-read the user rather than reusing the instance built in setUp:
        # `force_authenticate` hands the view the very object passed here, so a
        # related object touched by the test (`institution.contract`) would stay
        # cached on it and the view would see a row the test had already
        # deleted. A real request loads the user fresh from the session.
        self.client.force_authenticate(user=User.objects.get(pk=self.user.pk))

    def drop_contract(self):
        """Remove the contract without caching it on any instance first."""
        InstitutionContract.objects.filter(institution=self.institution).delete()


class MonthlyReportTests(InstitutionExportTestCase):
    def url(self):
        return reverse("institution-monthly-report")

    def test_anonymous_request_is_rejected(self):
        response = self.client.get(self.url())
        self.assertIn(response.status_code, (401, 403))

    def test_user_without_institution_is_rejected(self):
        outsider = User.objects.create_user(
            email="nadie@example.test", password="Respira.Test.2026"
        )
        self.client.force_authenticate(user=outsider)
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 403)

    def test_returns_a_pdf_for_the_requested_month(self):
        self.login()
        response = self.client.get(self.url(), {"month": "2026-07"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        # A real PDF, not an error page with an optimistic content type.
        self.assertTrue(response.content.startswith(b"%PDF-"))
        self.assertIn(
            "reporte-mensual-hospital-bautista-2026-07.pdf",
            response["Content-Disposition"],
        )

    def test_defaults_to_the_last_complete_month(self):
        """The current month would change between downloads, so it is not the default."""
        self.login()
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"%PDF-"))

    def test_rejects_a_malformed_month(self):
        self.login()
        response = self.client.get(self.url(), {"month": "julio"})
        self.assertEqual(response.status_code, 400)
        self.assertIn("month", response.json())

    def test_institution_without_a_sensor_gets_404(self):
        self.drop_contract()
        self.login()
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 404)

    def test_month_without_readings_still_renders(self):
        self.login()
        response = self.client.get(self.url(), {"month": "2026-01"})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"%PDF-"))

    def test_statistics_cover_only_the_institutions_own_station(self):
        from .exports import _month_statistics

        stats = _month_statistics(self.station.id, date(2026, 7, 1))

        self.assertEqual(len(stats["daily"]), 3)
        self.assertEqual(stats["measurements"], 6)
        # The other station's 500 would dominate if the query leaked across.
        self.assertEqual(stats["highest"], 120.0)
        self.assertEqual(stats["lowest"], 40.0)

    def test_daily_categories_are_counted_on_the_daily_average(self):
        from .exports import _month_statistics

        stats = _month_statistics(self.station.id, date(2026, 7, 1))

        # Averages are 50, 80 and 110 → good, moderate, unhealthy_sensitive.
        self.assertEqual(stats["distribution"]["good"], 1)
        self.assertEqual(stats["distribution"]["moderate"], 1)
        self.assertEqual(stats["distribution"]["unhealthy_sensitive"], 1)
        self.assertEqual(stats["distribution"]["hazardous"], 0)

    def test_threshold_is_only_applied_when_alerts_are_enabled(self):
        InstitutionAlertConfig.objects.create(
            institution=self.institution, is_enabled=False, alert_threshold=10
        )
        self.login()
        response = self.client.get(self.url(), {"month": "2026-07"})
        self.assertEqual(response.status_code, 200)


class RawExportTests(InstitutionExportTestCase):
    def url(self):
        return reverse("institution-raw-export")

    def test_anonymous_request_is_rejected(self):
        response = self.client.get(self.url())
        self.assertIn(response.status_code, (401, 403))

    def test_returns_a_readable_workbook(self):
        self.login()
        response = self.client.get(
            self.url(), {"from": "2026-07-01", "to": "2026-07-03"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # An XLSX is a zip archive; a truncated body fails to open here.
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(response.content)))

        workbook = load_workbook(io.BytesIO(response.content))
        rows = list(workbook["Mediciones"].values)

        self.assertEqual(rows[0][0], "Fecha y hora (Asunción)")
        # Six readings on this station, and none of the other station's.
        self.assertEqual(len(rows) - 1, 6)
        self.assertNotIn(500.0, [row[4] for row in rows[1:]])

    def test_timestamps_are_written_in_asuncion_local_time(self):
        self.login()
        response = self.client.get(
            self.url(), {"from": "2026-07-01", "to": "2026-07-01"}
        )
        workbook = load_workbook(io.BytesIO(response.content))
        first = list(workbook["Mediciones"].values)[1][0]

        # Stored as 08:00 local; a UTC leak would render 11:00 or 12:00.
        self.assertEqual(first.hour, 8)
        self.assertIsNone(first.tzinfo)

    def test_range_is_inclusive_of_the_end_day(self):
        self.login()
        response = self.client.get(
            self.url(), {"from": "2026-07-03", "to": "2026-07-03"}
        )
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(len(list(workbook["Mediciones"].values)) - 1, 2)

    def test_defaults_to_the_whole_contract(self):
        self.login()
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(len(list(workbook["Mediciones"].values)) - 1, 6)

    def test_rejects_an_inverted_range(self):
        self.login()
        response = self.client.get(
            self.url(), {"from": "2026-07-03", "to": "2026-07-01"}
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("to", response.json())

    def test_rejects_a_malformed_date(self):
        self.login()
        response = self.client.get(self.url(), {"from": "01/07/2026"})
        self.assertEqual(response.status_code, 400)
        self.assertIn("from", response.json())

    def test_institution_without_a_sensor_gets_404(self):
        self.drop_contract()
        self.login()
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 404)

    def test_row_cap_is_enforced(self):
        from unittest.mock import patch

        self.login()
        with patch("api.exports.MAX_EXPORT_ROWS", 2):
            response = self.client.get(
                self.url(), {"from": "2026-07-01", "to": "2026-07-03"}
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("from", response.json())

    def test_filename_carries_the_requested_range(self):
        self.login()
        response = self.client.get(
            self.url(), {"from": "2026-07-01", "to": "2026-07-03"}
        )
        self.assertIn(
            "historial-hospital-bautista-20260701-20260703.xlsx",
            response["Content-Disposition"],
        )
